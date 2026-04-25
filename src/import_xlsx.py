"""
Import PLA2 domains from Hull et al. supplemental xlsx.

Reads sequences from the "Plasmid Library > Passage P0" sheet, strips
the AAV9 homology arms (9 aa prefix + 10 aa suffix) to extract the
81-aa core PLA2 domain, and writes a FASTA file containing:
  - AAV9 wild type (first entry)
  - All natural parvoviral PLA2 domains

Excluded: synthetic sequences, disulfide variants (AAV9C##C##),
and active site mutants (D->A/N or H->A/N substitutions).
"""

import re
import openpyxl

XLSX_PATH = "jvi.00799-25-s0001.xlsx"
OUTPUT_FASTA = "data/natural_pla2_domains.fasta"

PREFIX_LEN = 9  # upstream homology arm: QDNARGLVL
SUFFIX_LEN = 10  # downstream homology arm: LGLVEEAAKT
EXPECTED_CORE_LEN = 81


def classify_variant(sid):
    """Classify a variant ID for inclusion/exclusion.

    Returns:
        "aav9_wt"    - AAV9 wild type reference (include as first entry)
        "synthetic"  - synthetic control sequence (exclude)
        "disulfide"  - disulfide bond variant AAV9C##C## (exclude)
        "as_mutant"  - active site mutant D/H -> A/N (exclude)
        "natural"    - natural parvoviral PLA2 domain (include)
    """
    sl = sid.lower()
    if "aav9(aas" in sl:
        return "aav9_wt"
    if sl.startswith("syntheticseq"):
        return "synthetic"
    if re.match(r"aav9c\d", sl):
        return "disulfide"
    if re.match(r"aav9[dh]\d+[an]$", sl):
        return "as_mutant"
    return "natural"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Plasmid Library > Passage P0"]

    natural = []
    aav9_entry = None
    excluded_counts = {"synthetic": 0, "disulfide": 0, "as_mutant": 0}

    for r in range(2, ws.max_row + 1):
        sid = str(ws.cell(r, 1).value or "").strip()
        seq = str(ws.cell(r, 2).value or "").strip()
        if not sid or not seq:
            continue

        core = seq[PREFIX_LEN:-SUFFIX_LEN]
        if len(core) != EXPECTED_CORE_LEN:
            print(f"  WARNING: {sid} core length {len(core)} != {EXPECTED_CORE_LEN}, skipping")
            continue

        variant_type = classify_variant(sid)
        if variant_type == "aav9_wt":
            aav9_entry = (sid, core)
        elif variant_type == "natural":
            natural.append((sid, core))
        else:
            excluded_counts[variant_type] += 1

    print(f"Excluded: {excluded_counts}")

    # AAV9 WT goes first
    all_entries = []
    if aav9_entry:
        all_entries.append(aav9_entry)
    all_entries.extend(natural)

    print(f"Natural PLA2 domains (including AAV9 WT): {len(all_entries)}")

    # Verify DxxxxxHD motif presence
    n_with_motif = sum(1 for _, s in all_entries if re.search(r"D.{5}HD", s))
    print(f"With DxxxxxHD motif: {n_with_motif}/{len(all_entries)}")

    with open(OUTPUT_FASTA, "w") as f:
        for sid, core in all_entries:
            f.write(f">{sid}\n{core}\n")
    print(f"Wrote {len(all_entries)} sequences to {OUTPUT_FASTA}")


if __name__ == "__main__":
    main()
